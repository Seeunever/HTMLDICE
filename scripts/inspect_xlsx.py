#!/usr/bin/env python3
import json
import sys
from pathlib import Path

import openpyxl

FILES = [
    ("莉莉娅", Path(r"e:/trpg/coc7/莉莉娅.xlsx")),
    ("特朗狗", Path(r"e:/trpg/coc7/特朗狗.xlsx")),
    ("奥八嘎", Path(r"e:/trpg/coc7/奥八嘎.xlsx")),
]


def cell_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text in ("☐", "☑", "□", "■"):
        return ""
    return text


def inspect_sheet(ws, max_rows=200, max_cols=30):
    cells = []
    for r in range(1, min(max_rows, ws.max_row + 1)):
        for c in range(1, min(max_cols, ws.max_column + 1)):
            text = cell_text(ws.cell(r, c).value)
            if text:
                cells.append({"r": r, "c": c, "v": text})
    return cells


def main():
    out = {}
    for username, path in FILES:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        out[username] = {
            "file": path.name,
            "sheet": ws.title,
            "size": [ws.max_row, ws.max_column],
            "cells": inspect_sheet(ws),
        }
    Path("e:/trpg/HTMLDICE/scripts/inspect_output.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print("done", len(out))


if __name__ == "__main__":
    main()
