"""CoC7 标准人物卡 xlsx 解析与格式校验（与莉莉娅人物卡模板一致）。"""

import re
import tempfile
from pathlib import Path

import openpyxl


def text(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s in ("☐", "☑", "□", "■", "√"):
        return s if s == "√" else ""
    return s


def number(value):
    if value is None or value == "":
        return None
    try:
        if isinstance(value, float) and value.is_integer():
            return int(value)
        if isinstance(value, int):
            return value
        s = str(value).strip().replace("%", "")
        if re.fullmatch(r"-?\d+", s):
            return int(s)
        if re.fullmatch(r"-?\d+\.\d+", s):
            return float(s)
    except (TypeError, ValueError):
        pass
    return None


def cell(ws, row, col):
    return ws.cell(row, col).value


def parse_skill_block(ws, row, name_col, value_col, occ_col):
    name = text(cell(ws, row, name_col))
    if not name or name in ("技能名称", "成功标", "本职"):
        return None
    value = number(cell(ws, row, value_col))
    if value is None or value <= 0:
        return None
    occ = text(cell(ws, row, occ_col))
    return {
        "name": name,
        "value": value,
        "occupation": occ == "★",
    }


def parse_skills(ws):
    skills = []
    for row in range(16, 95):
        left = parse_skill_block(ws, row, 6, 18, 4)
        if left:
            skills.append(left)
        right = parse_skill_block(ws, row, 28, 40, 26)
        if right:
            skills.append(right)
    skills.sort(key=lambda s: (-s["value"], s["name"]))
    return skills


def parse_weapons(ws):
    weapons = []
    for row in range(52, 60):
        name = text(cell(ws, row, 2))
        detail = text(cell(ws, row, 7))
        if not name or name.startswith("←") or name in ("武器", "武器名"):
            continue
        if not detail and not number(cell(ws, row, 17)):
            continue
        weapons.append(
            {
                "name": name,
                "detail": detail,
                "skill": text(cell(ws, row, 13)),
                "damage": text(cell(ws, row, 23)),
                "range": text(cell(ws, row, 17)),
                "attacks": text(cell(ws, row, 19)),
                "ammo": text(cell(ws, row, 21)),
            }
        )
    return weapons


def parse_assets(ws):
    return {
        "credit_rating": text(cell(ws, 62, 2)),
        "living_standard": text(cell(ws, 62, 6)),
        "spending_level": text(cell(ws, 62, 9)),
        "other_assets": text(cell(ws, 62, 12)),
        "cash": text(cell(ws, 62, 15)),
        "unit": text(cell(ws, 62, 19)),
        "description": text(cell(ws, 63, 2)),
        "asset_detail": text(cell(ws, 63, 12)),
    }


def parse_background(ws):
    rows = []
    for row in range(61, 95):
        label = text(cell(ws, row, 23))
        value = text(cell(ws, row, 27))
        if not label or label in ("背景故事", "形象描述"):
            continue
        if not value:
            continue
        if label.startswith("例：") and value.startswith("例："):
            continue
        rows.append({"label": label, "value": value})
    return rows


def calc_mov(str_val, dex_val, siz_val):
    if str_val is None or dex_val is None or siz_val is None:
        return None
    if dex_val <= siz_val and str_val <= siz_val:
        return 7
    if dex_val > siz_val and str_val > siz_val:
        return 9
    if dex_val > siz_val or str_val > siz_val:
        return 8
    return 7


def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=True)


def validate_character_xlsx(path):
    """校验 xlsx 是否符合莉莉娅同款 CoC7 人物卡布局。"""
    errors = []
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        return False, ["请上传 .xlsx 格式的 Excel 人物卡"]

    try:
        wb = load_workbook(path)
    except Exception:
        return False, ["无法读取 Excel 文件，请确认文件未损坏且为 .xlsx 格式"]

    if not wb.sheetnames:
        return False, ["Excel 文件中没有工作表"]

    ws = wb[wb.sheetnames[0]]

    checks = [
        ("调查员姓名", lambda: text(cell(ws, 3, 5))),
        ("职业", lambda: text(cell(ws, 5, 5))),
        ("力量 STR", lambda: number(cell(ws, 3, 21))),
        ("敏捷 DEX", lambda: number(cell(ws, 3, 27))),
        ("体质 CON", lambda: number(cell(ws, 5, 21))),
        ("意志 POW", lambda: number(cell(ws, 3, 33))),
        ("HP", lambda: number(cell(ws, 10, 7))),
        ("SAN", lambda: number(cell(ws, 10, 16))),
    ]

    for label, getter in checks:
        value = getter()
        if value is None or value == "":
            errors.append(f"缺少必填项：{label}（请使用与莉莉娅人物卡相同的标准模板）")

    skills = parse_skills(ws)
    if len(skills) < 3:
        errors.append("未识别到足够的技能数据，请使用标准 CoC7 人物卡模板（技能区布局需与莉莉娅人物卡一致）")

    attrs = [
        number(cell(ws, 3, 21)),
        number(cell(ws, 3, 27)),
        number(cell(ws, 5, 21)),
        number(cell(ws, 7, 21)),
        number(cell(ws, 3, 33)),
    ]
    if sum(1 for a in attrs if a is not None) < 4:
        errors.append("属性区格式不正确，请使用与莉莉娅人物卡相同的属性布局")

    return len(errors) == 0, errors


def parse_character_xlsx(path, owner, pc_id):
    """从 xlsx 解析人物卡数据。"""
    ok, errors = validate_character_xlsx(path)
    if not ok:
        raise ValueError("; ".join(errors))

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    attributes = [
        {"key": "STR", "label": "力量", "value": number(cell(ws, 3, 21)), "half": number(cell(ws, 3, 23))},
        {"key": "DEX", "label": "敏捷", "value": number(cell(ws, 3, 27)), "half": number(cell(ws, 3, 29))},
        {"key": "CON", "label": "体质", "value": number(cell(ws, 5, 21)), "half": number(cell(ws, 5, 23))},
        {"key": "APP", "label": "外貌", "value": number(cell(ws, 5, 27)), "half": number(cell(ws, 5, 29))},
        {"key": "SIZ", "label": "体型", "value": number(cell(ws, 7, 21)), "half": number(cell(ws, 7, 23))},
        {"key": "INT", "label": "智力", "value": number(cell(ws, 7, 27)), "half": number(cell(ws, 7, 29))},
        {"key": "POW", "label": "意志", "value": number(cell(ws, 3, 33)), "half": number(cell(ws, 3, 35))},
        {"key": "EDU", "label": "教育", "value": number(cell(ws, 5, 33)), "half": number(cell(ws, 5, 35))},
        {"key": "LUK", "label": "幸运", "value": number(cell(ws, 7, 33)), "half": number(cell(ws, 7, 35))},
    ]

    mov = calc_mov(attributes[0]["value"], attributes[1]["value"], attributes[4]["value"])
    hp = number(cell(ws, 10, 7))
    san = number(cell(ws, 10, 16))

    return {
        "pc_id": pc_id,
        "owner": owner,
        "profile": {
            "name": text(cell(ws, 3, 5)),
            "player": text(cell(ws, 4, 5)),
            "era": text(cell(ws, 4, 13)),
            "occupation": text(cell(ws, 5, 5)),
            "occupation_id": text(cell(ws, 5, 13)),
            "age": number(cell(ws, 6, 5)),
            "gender": text(cell(ws, 6, 13)),
            "residence": text(cell(ws, 7, 5)),
            "birthplace": text(cell(ws, 7, 13)),
            "date": " ".join(
                x
                for x in [
                    text(cell(ws, 8, 5)),
                    text(cell(ws, 8, 7)),
                    text(cell(ws, 8, 10)),
                    text(cell(ws, 8, 12)),
                    text(cell(ws, 8, 14)),
                ]
                if x
            ),
        },
        "status": {
            "hp": hp,
            "hp_max": hp,
            "hp_state": text(cell(ws, 11, 9)),
            "san": san,
            "san_max": max(san or 0, 99),
            "san_state": text(cell(ws, 11, 18)),
            "mp": number(cell(ws, 10, 25)),
            "mov": mov,
            "major_wound": number(cell(ws, 12, 4)),
            "temp_hp": number(cell(ws, 12, 9)),
            "san_loss_today": number(cell(ws, 12, 14)),
            "san_remaining": number(cell(ws, 12, 18)),
        },
        "occupation_note": text(cell(ws, 13, 2)),
        "attributes": attributes,
        "skills": parse_skills(ws),
        "weapons": parse_weapons(ws),
        "assets": parse_assets(ws),
        "appearance": text(cell(ws, 61, 27)),
        "background": parse_background(ws),
    }


def parse_character_bytes(file_bytes, owner, pc_id, filename="upload.xlsx"):
    suffix = Path(filename).suffix.lower() or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        return parse_character_xlsx(tmp_path, owner, pc_id)
    finally:
        Path(tmp_path).unlink(missing_ok=True)
